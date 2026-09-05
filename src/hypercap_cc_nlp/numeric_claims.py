"""Provenance-aware consistency checks for repeated aggregate numeric claims."""

from __future__ import annotations

import csv
import fnmatch
import json
import math
import re
import shutil
import subprocess
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCHEMA_VERSION = 1
REPORT_YAML_NAME = "NUMERIC_CONSISTENCY_REPORT.yml"
REPORT_MARKDOWN_NAME = "NUMERIC_CONSISTENCY_REPORT.md"
_ERROR_SEVERITIES = {"fatal", "major", "moderate"}
_PDF_N_RE = re.compile(r"[Nn]\s*=\s*([0-9][0-9,]*)")


class NumericClaimsError(ValueError):
    """Raised when the claim ledger or an evidence source is malformed."""


def load_ledger(path: Path) -> dict[str, Any]:
    """Load the JSON-compatible YAML ledger and validate its core schema."""

    try:
        ledger = json.loads(path.read_text())
    except (OSError, json.JSONDecodeError) as exc:
        raise NumericClaimsError(f"Cannot load numeric claim ledger {path}: {exc}") from exc
    if ledger.get("schema_version") != SCHEMA_VERSION:
        raise NumericClaimsError(
            f"Unsupported numeric claim schema: {ledger.get('schema_version')!r}; "
            f"expected {SCHEMA_VERSION}."
        )
    metadata = ledger.get("metadata")
    claims = ledger.get("claims")
    if not isinstance(metadata, dict) or not isinstance(claims, list):
        raise NumericClaimsError("Ledger requires metadata mapping and claims list.")
    accepted_run = metadata.get("accepted_run", {})
    producer_sha = str(accepted_run.get("producer_sha", ""))
    if not re.fullmatch(r"[0-9a-f]{40}", producer_sha):
        raise NumericClaimsError("accepted_run.producer_sha must be a lowercase 40-character SHA.")
    claim_ids = [claim.get("claim_id") for claim in claims if isinstance(claim, dict)]
    if len(claim_ids) != len(claims) or any(not claim_id for claim_id in claim_ids):
        raise NumericClaimsError("Every claim requires a nonempty claim_id.")
    if len(set(claim_ids)) != len(claim_ids):
        raise NumericClaimsError("claim_id values must be unique.")
    for claim in claims:
        if "exact_value" not in claim or "source" not in claim:
            raise NumericClaimsError(f"Claim {claim['claim_id']} requires exact_value and source.")
        source = claim["source"]
        if source.get("base") not in {"repo", "results", "artifacts"}:
            raise NumericClaimsError(f"Claim {claim['claim_id']} has an invalid source base.")
        for target in claim.get("targets", []):
            if target.get("base") not in {"repo", "results", "artifacts"}:
                raise NumericClaimsError(f"Claim {claim['claim_id']} has an invalid target base.")
    return ledger


def _finding(
    *,
    claim_id: str,
    severity: str,
    issue_type: str,
    location: str,
    diagnosis: str,
    recommended_fix: str,
) -> dict[str, str]:
    return {
        "claim_id": claim_id,
        "severity": severity,
        "confidence": "high",
        "issue_type": issue_type,
        "location": location,
        "diagnosis": diagnosis,
        "recommended_fix": recommended_fix,
    }


def _matches(candidate: Mapping[str, Any], expected: Mapping[str, Any]) -> bool:
    return all(candidate.get(key) == value for key, value in expected.items())


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle))


def _xlsx_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise NumericClaimsError(f"Missing sheet {sheet_name!r} in {path}.")
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        raise NumericClaimsError(f"Empty sheet {sheet_name!r} in {path}.")
    headers = list(rows[0])
    if any(header is None for header in headers) or len(set(headers)) != len(headers):
        raise NumericClaimsError(f"Invalid or duplicate headers in {path}::{sheet_name}.")
    return [dict(zip(headers, row, strict=True)) for row in rows[1:] if any(value is not None for value in row)]


def _json_path_value(payload: Any, keys: Sequence[str]) -> Any:
    value = payload
    for key in keys:
        if not isinstance(value, dict) or key not in value:
            raise NumericClaimsError(f"Missing JSON path component {key!r}.")
        value = value[key]
    return value


def _base_root(
    base: str,
    *,
    repo_root: Path,
    results_dir: Path | None,
    artifacts_dir: Path | None,
) -> Path | None:
    return {"repo": repo_root, "results": results_dir, "artifacts": artifacts_dir}[base]


def _source_value(
    source: Mapping[str, Any],
    *,
    repo_root: Path,
    results_dir: Path | None,
    artifacts_dir: Path | None,
) -> tuple[Any, str]:
    base = str(source["base"])
    root = _base_root(
        base,
        repo_root=repo_root,
        results_dir=results_dir,
        artifacts_dir=artifacts_dir,
    )
    if root is None:
        raise FileNotFoundError(f"No {base} directory was supplied.")
    path = root / str(source["path"])
    if not path.is_file():
        raise FileNotFoundError(path)
    kind = source["kind"]
    if kind == "json_path":
        value = _json_path_value(json.loads(path.read_text()), source["json_path"])
        return value, f"{path}::{'.'.join(source['json_path'])}"
    if kind == "csv_row":
        rows = [row for row in _csv_rows(path) if _matches(row, source["match"])]
        if len(rows) != 1:
            raise NumericClaimsError(f"Expected one matching row in {path}; found {len(rows)}.")
        column = source["value_column"]
        if column not in rows[0]:
            raise NumericClaimsError(f"Missing column {column!r} in {path}.")
        return _coerce_csv_value(rows[0][column]), f"{path}::{source['match']}::{column}"
    if kind in {"xlsx_row", "xlsx_sum"}:
        sheet = str(source["sheet"])
        rows = [row for row in _xlsx_rows(path, sheet) if _matches(row, source["match"])]
        column = source["value_column"]
        if kind == "xlsx_row":
            if len(rows) != 1:
                raise NumericClaimsError(
                    f"Expected one matching row in {path}::{sheet}; found {len(rows)}."
                )
            if column not in rows[0]:
                raise NumericClaimsError(f"Missing column {column!r} in {path}::{sheet}.")
            return rows[0][column], f"{path}::{sheet}::{source['match']}::{column}"
        if not rows:
            raise NumericClaimsError(f"No rows matched in {path}::{sheet}.")
        values = [row.get(column) for row in rows]
        if any(not isinstance(value, (int, float)) for value in values):
            raise NumericClaimsError(f"Cannot sum nonnumeric {column!r} values in {path}::{sheet}.")
        return sum(values), f"{path}::{sheet}::{source['match']}::sum({column})"
    raise NumericClaimsError(f"Unsupported source kind: {kind!r}.")


def _coerce_csv_value(value: str) -> Any:
    stripped = value.strip()
    if not stripped:
        return stripped
    if re.fullmatch(r"[-+]?\d+", stripped):
        return int(stripped)
    try:
        return float(stripped)
    except ValueError:
        return value


def _values_equal(expected: Any, observed: Any, tolerance: float) -> bool:
    if isinstance(expected, bool) or isinstance(observed, bool):
        return expected == observed
    if isinstance(expected, (int, float)) and isinstance(observed, (int, float)):
        if not math.isfinite(float(observed)):
            return False
        return math.isclose(float(expected), float(observed), rel_tol=0.0, abs_tol=tolerance)
    return expected == observed


def _target_findings(
    claim_id: str,
    target: Mapping[str, Any],
    *,
    repo_root: Path,
    results_dir: Path | None,
    artifacts_dir: Path | None,
) -> list[dict[str, str]]:
    base = str(target["base"])
    root = _base_root(
        base,
        repo_root=repo_root,
        results_dir=results_dir,
        artifacts_dir=artifacts_dir,
    )
    if root is None:
        return []
    path = root / str(target["path"])
    if not path.is_file():
        return [
            _finding(
                claim_id=claim_id,
                severity="moderate",
                issue_type="missing_in_expected_section",
                location=str(path),
                diagnosis="Expected active target is missing.",
                recommended_fix="Restore the expected target or remove it from the ledger with documented rationale.",
            )
        ]
    findings: list[dict[str, str]] = []
    kind = target["kind"]
    if kind == "text":
        text = path.read_text()
        for token in target.get("required", []):
            if token not in text:
                findings.append(
                    _finding(
                        claim_id=claim_id,
                        severity="moderate",
                        issue_type="missing_in_expected_section",
                        location=str(path),
                        diagnosis=f"Required canonical token is missing: {token!r}.",
                        recommended_fix="Restore the canonical target text from the numeric claim ledger.",
                    )
                )
        for token in target.get("forbidden", []):
            if token in text:
                findings.append(
                    _finding(
                        claim_id=claim_id,
                        severity="major",
                        issue_type="outdated_section",
                        location=str(path),
                        diagnosis=f"Stale token remains: {token!r}.",
                        recommended_fix="Replace the stale value with the source-backed ledger value.",
                    )
                )
        return findings
    if kind == "csv_row":
        rows = [row for row in _csv_rows(path) if _matches(row, target["match"])]
        if len(rows) != 1:
            return [
                _finding(
                    claim_id=claim_id,
                    severity="moderate",
                    issue_type="missing_in_expected_section",
                    location=str(path),
                    diagnosis=f"Expected one target row; found {len(rows)}.",
                    recommended_fix="Restore one uniquely keyed dictionary row.",
                )
            ]
        column = target["value_column"]
        observed = rows[0].get(column)
        expected = target["expected"]
        if observed != expected:
            findings.append(
                _finding(
                    claim_id=claim_id,
                    severity="major",
                    issue_type="outdated_section",
                    location=f"{path}::{target['match']}::{column}",
                    diagnosis=f"Observed {observed!r}; expected {expected!r}.",
                    recommended_fix="Replace the stale field with the exact ledger-approved wording.",
                )
            )
        return findings
    raise NumericClaimsError(f"Unsupported target kind: {kind!r}.")


def _is_excluded(relative_path: str, excluded_globs: Iterable[str]) -> bool:
    return any(fnmatch.fnmatch(relative_path, pattern) for pattern in excluded_globs)


def _active_files(repo_root: Path, metadata: Mapping[str, Any]) -> list[Path]:
    excluded = metadata.get("excluded_globs", [])
    files: set[Path] = set()
    for pattern in metadata.get("active_globs", []):
        for path in repo_root.glob(pattern):
            if path.is_file():
                relative = path.relative_to(repo_root).as_posix()
                if not _is_excluded(relative, excluded):
                    files.add(path)
    return sorted(files)


def audit_static_targets(
    ledger: Mapping[str, Any],
    *,
    repo_root: Path,
) -> list[dict[str, str]]:
    """Audit tracked/current narrative targets without requiring local results."""

    findings: list[dict[str, str]] = []
    for claim in ledger["claims"]:
        for target in claim.get("targets", []):
            if target["base"] != "repo":
                continue
            findings.extend(
                _target_findings(
                    claim["claim_id"],
                    target,
                    repo_root=repo_root,
                    results_dir=None,
                    artifacts_dir=None,
                )
            )
    stale_phrases = ledger["metadata"].get("stale_phrases", [])
    for path in _active_files(repo_root, ledger["metadata"]):
        try:
            text = path.read_text()
        except UnicodeDecodeError:
            continue
        for phrase in stale_phrases:
            if phrase in text:
                findings.append(
                    _finding(
                        claim_id="global_stale_phrase_scan",
                        severity="major",
                        issue_type="outdated_section",
                        location=str(path),
                        diagnosis=f"Known stale phrase remains: {phrase!r}.",
                        recommended_fix="Replace it with the accepted-run value and add a claim target if repeated.",
                    )
                )
    return findings


def audit_live_sources(
    ledger: Mapping[str, Any],
    *,
    repo_root: Path,
    results_dir: Path,
    artifacts_dir: Path,
    require_sources: bool,
) -> tuple[list[dict[str, str]], dict[str, dict[str, Any]]]:
    """Validate ledger claims and generated targets against aggregate run artifacts."""

    findings: list[dict[str, str]] = []
    observations: dict[str, dict[str, Any]] = {}
    for claim in ledger["claims"]:
        claim_id = claim["claim_id"]
        try:
            observed, location = _source_value(
                claim["source"],
                repo_root=repo_root,
                results_dir=results_dir,
                artifacts_dir=artifacts_dir,
            )
        except (FileNotFoundError, NumericClaimsError, OSError, json.JSONDecodeError) as exc:
            observations[claim_id] = {"status": "source_missing", "error": str(exc)}
            if require_sources:
                findings.append(
                    _finding(
                        claim_id=claim_id,
                        severity="major",
                        issue_type="missing_in_expected_section",
                        location=str(claim["source"].get("path", "")),
                        diagnosis=f"Canonical aggregate source could not be read: {exc}",
                        recommended_fix="Restore the accepted aggregate source; do not substitute an older run.",
                    )
                )
            continue
        expected = claim["exact_value"]
        tolerance = float(claim.get("tolerance", 0.0))
        matches = _values_equal(expected, observed, tolerance)
        observations[claim_id] = {
            "status": "consistent" if matches else "contradiction",
            "expected": expected,
            "observed": observed,
            "source": location,
            "tolerance": tolerance,
        }
        if not matches:
            findings.append(
                _finding(
                    claim_id=claim_id,
                    severity="major",
                    issue_type="contradiction",
                    location=location,
                    diagnosis=f"Aggregate source value {observed!r} does not match ledger {expected!r}.",
                    recommended_fix="Investigate the producer and require a new accepted run; do not edit the preserved run.",
                )
            )
        for target in claim.get("targets", []):
            if target["base"] == "repo":
                continue
            findings.extend(
                _target_findings(
                    claim_id,
                    target,
                    repo_root=repo_root,
                    results_dir=results_dir,
                    artifacts_dir=artifacts_dir,
                )
            )
    return findings, observations


def audit_figure_manifest(
    ledger: Mapping[str, Any],
    *,
    results_dir: Path,
) -> list[dict[str, str]]:
    path = results_dir / "figure_manifest.csv"
    if not path.is_file():
        return [
            _finding(
                claim_id="figure_manifest_denominators",
                severity="major",
                issue_type="missing_in_expected_section",
                location=str(path),
                diagnosis="Accepted figure manifest is missing.",
                recommended_fix="Restore the accepted manifest or regenerate a new dated run.",
            )
        ]
    expected = ledger.get("figure_manifest_denominators", {})
    observed: dict[str, int] = {}
    for row in _csv_rows(path):
        raw = row.get("cohort_denominator_n", "").strip()
        if raw:
            observed[row["figure_key"]] = int(raw)
    if observed == expected:
        return []
    missing = sorted(set(expected) - set(observed))
    extra = sorted(set(observed) - set(expected))
    mismatched = {
        key: {"expected": expected[key], "observed": observed[key]}
        for key in sorted(set(expected) & set(observed))
        if expected[key] != observed[key]
    }
    return [
        _finding(
            claim_id="figure_manifest_denominators",
            severity="major",
            issue_type="contradiction",
            location=str(path),
            diagnosis=f"Figure denominator mapping drifted; missing={missing}, extra={extra}, mismatched={mismatched}.",
            recommended_fix="Reconcile the producing notebook and create a new accepted run; do not patch the manifest.",
        )
    ]


def audit_accepted_provenance(
    ledger: Mapping[str, Any],
    *,
    results_dir: Path,
    artifacts_dir: Path,
) -> list[dict[str, str]]:
    accepted = ledger["metadata"]["accepted_run"]
    path = artifacts_dir / accepted["acceptance_audit"]
    if not path.is_file():
        return [
            _finding(
                claim_id="accepted_run_provenance",
                severity="major",
                issue_type="missing_in_expected_section",
                location=str(path),
                diagnosis="Sealed acceptance audit is missing.",
                recommended_fix="Restore the sealed acceptance evidence; do not infer acceptance from filenames.",
            )
        ]
    try:
        payload = json.loads(path.read_text())
    except json.JSONDecodeError as exc:
        return [
            _finding(
                claim_id="accepted_run_provenance",
                severity="major",
                issue_type="contradiction",
                location=str(path),
                diagnosis=f"Acceptance audit is not valid JSON: {exc}",
                recommended_fix="Restore the sealed acceptance evidence from its verified copy.",
            )
        ]
    expected = {
        "status": "pass",
        "code_sha": accepted["producer_sha"],
        "results_date": accepted["results_date"],
    }
    observed = {key: payload.get(key) for key in expected}
    findings: list[dict[str, str]] = []
    if observed != expected:
        findings.append(
            _finding(
                claim_id="accepted_run_provenance",
                severity="major",
                issue_type="contradiction",
                location=str(path),
                diagnosis=f"Acceptance identity mismatch: observed={observed}, expected={expected}.",
                recommended_fix="Use the exact accepted run and producer SHA; do not relabel another run.",
            )
        )
    if results_dir.name != accepted["results_date"]:
        findings.append(
            _finding(
                claim_id="accepted_run_provenance",
                severity="major",
                issue_type="contradiction",
                location=str(results_dir),
                diagnosis=(
                    f"Results directory name {results_dir.name!r} does not match accepted date "
                    f"{accepted['results_date']!r}."
                ),
                recommended_fix="Point live mode to the exact accepted Results directory.",
            )
        )
    return findings


def audit_submission_manifest(
    ledger: Mapping[str, Any],
    *,
    results_dir: Path,
) -> tuple[list[dict[str, str]], int]:
    path = results_dir / "submission_manifest.csv"
    if not path.is_file():
        return [
            _finding(
                claim_id="submission_manifest",
                severity="major",
                issue_type="missing_in_expected_section",
                location=str(path),
                diagnosis="Accepted submission manifest is missing.",
                recommended_fix="Restore the accepted manifest or regenerate a new dated run.",
            )
        ], 0
    expected_date = ledger["metadata"]["accepted_run"]["results_date"]
    manuscript_denominators: dict[str, int] = {}
    for key, value in ledger.get("figure_manifest_denominators", {}).items():
        match = re.fullmatch(r"figure_(s?\d+)_(?:pdf|png)", key)
        if match:
            suffix = match.group(1).upper()
            manuscript_denominators[f"Figure {suffix}"] = int(value)
    findings: list[dict[str, str]] = []
    pending_manual = 0
    for row in _csv_rows(path):
        if row.get("caption_check_status") == "pending_manual_check":
            pending_manual += 1
        if row.get("date_generated") != expected_date:
            findings.append(
                _finding(
                    claim_id="submission_manifest",
                    severity="major",
                    issue_type="outdated_section",
                    location=str(path),
                    diagnosis=f"{row.get('artifact_id')} has date {row.get('date_generated')!r}, expected {expected_date!r}.",
                    recommended_fix="Use only assets from the accepted dated run.",
                )
            )
        artifact_id = row.get("artifact_id", "")
        raw_denominator = row.get("cohort_denominator_n", "").strip()
        if artifact_id in manuscript_denominators and raw_denominator:
            observed = int(raw_denominator)
            expected = manuscript_denominators[artifact_id]
            if observed != expected:
                findings.append(
                    _finding(
                        claim_id="submission_manifest",
                        severity="major",
                        issue_type="contradiction",
                        location=f"{path}::{artifact_id}",
                        diagnosis=f"Submission denominator {observed} does not match figure manifest {expected}.",
                        recommended_fix="Correct the producer and generate a new dated submission manifest.",
                    )
                )
        final_path = row.get("final_file_path", "").strip()
        if final_path and not (results_dir / final_path).is_file():
            findings.append(
                _finding(
                    claim_id="submission_manifest",
                    severity="major",
                    issue_type="missing_in_expected_section",
                    location=str(results_dir / final_path),
                    diagnosis=f"Manifest-listed asset for {artifact_id!r} is missing.",
                    recommended_fix="Restore the accepted asset or generate a new dated submission package.",
                )
            )
    return findings, pending_manual


def extract_pdf_n_tokens(path: Path, *, pdftotext: str = "pdftotext") -> list[int]:
    """Extract unique N=/n= sample-size tokens from a PDF."""

    completed = subprocess.run(
        [pdftotext, "-layout", str(path), "-"],
        check=True,
        capture_output=True,
        text=True,
    )
    return sorted({int(token.replace(",", "")) for token in _PDF_N_RE.findall(completed.stdout)})


def audit_pdf_tokens(
    ledger: Mapping[str, Any],
    *,
    results_dir: Path,
    require_sources: bool,
) -> tuple[list[dict[str, str]], dict[str, list[int]]]:
    expected_mapping = ledger.get("pdf_n_tokens", {})
    if not expected_mapping:
        return [], {}
    executable = shutil.which("pdftotext")
    if executable is None:
        if not require_sources:
            return [], {}
        return [
            _finding(
                claim_id="publication_figure_pdf_tokens",
                severity="moderate",
                issue_type="missing_in_expected_section",
                location="pdftotext",
                diagnosis="pdftotext is required for live publication-figure token checks.",
                recommended_fix="Install Poppler/pdftotext or run static mode only.",
            )
        ], {}
    findings: list[dict[str, str]] = []
    observed_mapping: dict[str, list[int]] = {}
    for filename, expected in expected_mapping.items():
        path = results_dir / filename
        if not path.is_file():
            findings.append(
                _finding(
                    claim_id="publication_figure_pdf_tokens",
                    severity="major",
                    issue_type="missing_in_expected_section",
                    location=str(path),
                    diagnosis="Publication figure PDF is missing.",
                    recommended_fix="Restore the accepted figure or generate a new dated run.",
                )
            )
            continue
        try:
            observed = extract_pdf_n_tokens(path, pdftotext=executable)
        except subprocess.CalledProcessError as exc:
            findings.append(
                _finding(
                    claim_id="publication_figure_pdf_tokens",
                    severity="major",
                    issue_type="missing_in_expected_section",
                    location=str(path),
                    diagnosis=f"PDF text extraction failed: {exc}",
                    recommended_fix="Repair or regenerate the figure PDF before submission.",
                )
            )
            continue
        observed_mapping[filename] = observed
        if observed != expected:
            findings.append(
                _finding(
                    claim_id="publication_figure_pdf_tokens",
                    severity="major",
                    issue_type="contradiction",
                    location=str(path),
                    diagnosis=f"Extracted N tokens {observed} do not match accepted tokens {expected}.",
                    recommended_fix="Reconcile the figure source and generate a new dated run; do not patch the PDF.",
                )
            )
    return findings, observed_mapping


def build_report(
    ledger: Mapping[str, Any],
    *,
    findings: Sequence[Mapping[str, str]],
    observations: Mapping[str, Mapping[str, Any]],
    pdf_tokens: Mapping[str, Sequence[int]],
    pending_manual_caption_checks: int,
    live_mode: bool,
) -> dict[str, Any]:
    """Build the cross-document consistency matrix and concise audit summary."""

    findings_by_claim: dict[str, list[Mapping[str, str]]] = {}
    for finding in findings:
        findings_by_claim.setdefault(finding["claim_id"], []).append(finding)
    concepts: list[dict[str, Any]] = []
    for claim in ledger["claims"]:
        claim_id = claim["claim_id"]
        claim_findings = findings_by_claim.get(claim_id, [])
        if any(item["issue_type"] == "contradiction" for item in claim_findings):
            status = "contradiction"
        elif any(item["issue_type"] == "outdated_section" for item in claim_findings):
            status = "outdated_section"
        elif claim_findings:
            status = "missing_in_expected_section"
        else:
            status = "consistent"
        severity = max(
            (item["severity"] for item in claim_findings),
            key=lambda value: {"minor": 0, "moderate": 1, "major": 2, "fatal": 3}[value],
            default="minor",
        )
        source = claim["source"]
        mentions = [
            {
                "document": source["path"],
                "section": source.get("sheet", source.get("json_path", source.get("match", ""))),
                "value": observations.get(claim_id, {}).get("observed"),
            }
        ]
        mentions.extend(
            {
                "document": target["path"],
                "section": target.get("match", "active target"),
                "value": claim["exact_value"],
            }
            for target in claim.get("targets", [])
        )
        concepts.append(
            {
                "concept_id": claim_id,
                "concept_type": claim["concept_type"],
                "canonical_value": claim["exact_value"],
                "mentions": mentions,
                "consistency_status": status,
                "severity": severity,
                "recommended_resolution": (
                    "Retain the accepted-run value."
                    if status == "consistent"
                    else "; ".join(item["recommended_fix"] for item in claim_findings)
                ),
                "sections_to_update": sorted({item["location"] for item in claim_findings}),
            }
        )
    unassigned = [
        finding
        for finding in findings
        if finding["claim_id"] not in {claim["claim_id"] for claim in ledger["claims"]}
    ]
    failing = [finding for finding in findings if finding["severity"] in _ERROR_SEVERITIES]
    return {
        "summary": {
            "status": "fail" if failing else "pass",
            "mode": "live" if live_mode else "static",
            "claims": len(ledger["claims"]),
            "findings": len(findings),
            "failing_findings": len(failing),
            "publication_figure_pdfs_checked": len(pdf_tokens),
            "pending_manual_caption_checks_not_resolved_by_numeric_audit": pending_manual_caption_checks,
        },
        "consistency_matrix": {
            "metadata": {
                "package_type": ledger["metadata"]["package_type"],
                "venue_profile": ledger["metadata"]["venue_profile"],
                "accepted_run": ledger["metadata"]["accepted_run"],
                "mode": "live" if live_mode else "static",
            },
            "canonical_terms": [
                {
                    "concept_id": "scientific_execution",
                    "preferred_term": f"accepted run produced at {ledger['metadata']['accepted_run']['producer_sha']}",
                    "variants_found": [],
                    "recommended_action": "Attribute numerical results to the producing SHA, not the merge commit.",
                }
            ],
            "concepts": concepts,
            "high_risk_drift": [
                finding for finding in findings if finding["severity"] in {"fatal", "major"}
            ],
            "unassigned_findings": unassigned,
        },
        "source_observations": observations,
        "pdf_n_tokens": pdf_tokens,
        "excluded_surfaces": ledger["metadata"].get("excluded_globs", []),
    }


def render_report_markdown(report: Mapping[str, Any]) -> str:
    summary = report["summary"]
    lines = [
        "# Numeric Consistency Audit",
        "",
        f"- Status: **{summary['status'].upper()}**",
        f"- Mode: `{summary['mode']}`",
        f"- Canonical claims checked: {summary['claims']}",
        f"- Findings: {summary['findings']} ({summary['failing_findings']} failing)",
        f"- Publication figure PDFs checked: {summary['publication_figure_pdfs_checked']}",
        "- Pending manual caption/source checks outside this numeric audit: "
        f"{summary['pending_manual_caption_checks_not_resolved_by_numeric_audit']}",
        "",
        "## Claim status",
        "",
        "| Claim | Status | Canonical value | Severity |",
        "| --- | --- | ---: | --- |",
    ]
    for concept in report["consistency_matrix"]["concepts"]:
        lines.append(
            f"| `{concept['concept_id']}` | {concept['consistency_status']} | "
            f"{concept['canonical_value']} | {concept['severity']} |"
        )
    unassigned = report["consistency_matrix"]["unassigned_findings"]
    lines.extend(["", "## Other findings", ""])
    if not unassigned:
        lines.append("None.")
    else:
        for finding in unassigned:
            lines.append(
                f"- **{finding['severity']}** `{finding['issue_type']}` at "
                f"`{finding['location']}`: {finding['diagnosis']}"
            )
    lines.extend(
        [
            "",
            "## Boundary",
            "",
            "Manuscript/preprint files and immutable historical evidence were excluded. "
            "This report contains aggregate values and provenance only.",
            "",
        ]
    )
    return "\n".join(lines)


def write_reports(report: Mapping[str, Any], report_dir: Path) -> tuple[Path, Path]:
    report_dir.mkdir(parents=True, exist_ok=True)
    yaml_path = report_dir / REPORT_YAML_NAME
    markdown_path = report_dir / REPORT_MARKDOWN_NAME
    yaml_path.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n")
    markdown_path.write_text(render_report_markdown(report))
    return yaml_path, markdown_path


def run_audit(
    *,
    repo_root: Path,
    ledger_path: Path,
    results_dir: Path | None = None,
    artifacts_dir: Path | None = None,
    require_sources: bool = False,
) -> dict[str, Any]:
    ledger = load_ledger(ledger_path)
    findings = audit_static_targets(ledger, repo_root=repo_root)
    observations: dict[str, dict[str, Any]] = {}
    pdf_tokens: dict[str, list[int]] = {}
    pending_manual = 0
    live_mode = results_dir is not None
    if live_mode:
        if artifacts_dir is None:
            raise NumericClaimsError("Live mode requires artifacts_dir.")
        source_findings, observations = audit_live_sources(
            ledger,
            repo_root=repo_root,
            results_dir=results_dir,
            artifacts_dir=artifacts_dir,
            require_sources=require_sources,
        )
        findings.extend(source_findings)
        findings.extend(
            audit_accepted_provenance(
                ledger,
                results_dir=results_dir,
                artifacts_dir=artifacts_dir,
            )
        )
        findings.extend(audit_figure_manifest(ledger, results_dir=results_dir))
        submission_findings, pending_manual = audit_submission_manifest(
            ledger, results_dir=results_dir
        )
        findings.extend(submission_findings)
        pdf_findings, pdf_tokens = audit_pdf_tokens(
            ledger,
            results_dir=results_dir,
            require_sources=require_sources,
        )
        findings.extend(pdf_findings)
    return build_report(
        ledger,
        findings=findings,
        observations=observations,
        pdf_tokens=pdf_tokens,
        pending_manual_caption_checks=pending_manual,
        live_mode=live_mode,
    )
